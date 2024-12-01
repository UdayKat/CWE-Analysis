import json
from openpyxl import load_workbook

# Load JSON file
with open(r'C:\Users\udayk\OneDrive\Desktop\CWEAnalysis\MSR_data_cleaned.json', 'r') as f:
    data = json.load(f)

# Filter data
filtered_data = []
for number in data:
    if data[number]["CWE ID"] == "CWE-89" and data[number]["vul"] == '0':
        filtered_data.append(data[number]['func_before'])   
    if len(filtered_data) == 10:
        break 

# Load Excel workbook
wb = load_workbook("CWE-89_not_vul.xlsx")
ws = wb.active

# Clear existing content in the sheet (optional)
for row in ws.iter_rows():
    for cell in row:
        cell.value = None

# Fill the sheet with the filtered data
for index, value in enumerate(filtered_data, start=1):
    ws.cell(row=index, column=1, value=value)  # Writing data into the first column

# Save the updated workbook
wb.save("CWE-89_not_vul.xlsx")

print("Excel sheet updated successfully!")
